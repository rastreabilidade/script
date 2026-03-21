import pandas as pd
import matplotlib.pyplot as plt
import os
import re
import textwrap
import numpy as np
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

# 1. Configurações Iniciais
caminho_arquivo_entrada = 'Mapa final - Castanhal - Produtores.xlsx'
caminho_arquivo_saida = 'Testefinal.xlsx'
limite_caracteres_titulo = 75

# Lista para guardar o nome das imagens temporárias e apagá-las no final
imagens_temporarias = []
nomes_abas_usados = []

try:
    # Lendo todas as abas do Excel de uma vez
    dfs = pd.read_excel(caminho_arquivo_entrada, sheet_name=None)

    print(f"Lendo o arquivo '{caminho_arquivo_entrada}'...")
    print(f"Foram encontradas {len(dfs)} abas. Iniciando processamento...\n")

    # Inicializa o escritor do Excel
    with pd.ExcelWriter(caminho_arquivo_saida, engine='openpyxl') as writer:

        for nome_aba_original, df in dfs.items():
            # Se a aba estiver vazia, pula para a próxima
            if df.empty or len(df.columns) == 0:
                print(f"Aba '{nome_aba_original}' está vazia. Ignorando.")
                continue

            # Assumimos que a coluna a ser contada é a PRIMEIRA coluna da aba
            nome_coluna_alvo = df.columns[0]

            # --- Regras do Excel para Nomes de Abas ---
            novo_nome_aba = re.sub(r'[\\/*?:\[\]]', '', str(nome_coluna_alvo))
            novo_nome_aba = novo_nome_aba[:30]

            # Evita erro caso duas abas acabem com o exato mesmo nome
            if novo_nome_aba in nomes_abas_usados:
                novo_nome_aba = f"{novo_nome_aba}_{len(nomes_abas_usados)}"
            nomes_abas_usados.append(novo_nome_aba)

            print(f"Processando: '{nome_aba_original}' -> Criando aba '{novo_nome_aba}'")

            # 1. Escreve os dados originais na nova aba
            df.to_excel(writer, sheet_name=novo_nome_aba, index=False)

            # 2. Realiza a contagem e porcentagem
            contagem = df[nome_coluna_alvo].value_counts()
            porcentagem = df[nome_coluna_alvo].value_counts(normalize=True) * 100

            tabela_resumo = pd.DataFrame({
                'Quantidade': contagem,
                '%': porcentagem.round(2)
            })

            tabela_resumo.loc['Total'] = [contagem.sum(), porcentagem.sum().round(2)]

            # Calcula onde colocar a tabela de resumo
            coluna_inicio_resumo = df.shape[1] + 1

            # Escreve a tabela de resumo
            tabela_resumo.to_excel(
                writer,
                sheet_name=novo_nome_aba,
                startcol=coluna_inicio_resumo,
                index_label='Respostas'
            )

            # Remove a linha Total para os gráficos
            dados_grafico = tabela_resumo.drop('Total', errors='ignore')

            # Se não houver dados válidos, pula
            if dados_grafico.empty:
                print(f"Aba '{novo_nome_aba}' sem dados para gráficos. Ignorando gráficos.")
                continue

            total = dados_grafico['Quantidade'].sum()
            limite_percentual = 2.0
            titulo_formatado = textwrap.fill(str(nome_coluna_alvo), width=limite_caracteres_titulo)

            # =========================
            # 3. GRÁFICO DE PIZZA
            # =========================
            fig_pizza, ax_pizza = plt.subplots(figsize=(14, 8), subplot_kw=dict(aspect="equal"))

            def formato_porcentagem_dentro(pct):
                return ('%1.1f%%' % pct) if pct > limite_percentual else ''

            wedges, texts, autotexts = ax_pizza.pie(
                dados_grafico['Quantidade'],
                autopct=formato_porcentagem_dentro,
                startangle=140,
                colors=plt.cm.Set3.colors,
                pctdistance=0.75
            )

            # Rótulos externos ao redor da pizza
            for i, p in enumerate(wedges):
                valor = dados_grafico['Quantidade'].iloc[i]
                pct = (valor / total) * 100 if total != 0 else 0

                if pct <= limite_percentual and pct > 0:
                    ang = (p.theta2 - p.theta1) / 2.0 + p.theta1

                    x = np.cos(np.deg2rad(ang))
                    y = np.sin(np.deg2rad(ang))

                    x_seta = 1.00 * x
                    y_seta = 1.00 * y

                    x_text = 1.25 * x
                    y_text = 1.25 * y

                    ha = "left" if x >= 0 else "right"

                    ax_pizza.annotate(
                        f"{pct:.1f}%",
                        xy=(x_seta, y_seta),
                        xytext=(x_text, y_text),
                        horizontalalignment=ha,
                        fontsize=9,
                        va="center",
                        arrowprops=dict(
                            arrowstyle="-",
                            color="black",
                            lw=0.6
                        )
                    )

            numero_de_respostas = len(dados_grafico)
            max_por_coluna = 40
            colunas_dinamicas = (numero_de_respostas - 1) // max_por_coluna + 1

            ax_pizza.legend(
                wedges,
                dados_grafico.index,
                loc="center left",
                bbox_to_anchor=(1.15, 0.5),
                frameon=False,
                fontsize=9,
                ncol=colunas_dinamicas
            )

            ax_pizza.set_title(titulo_formatado, fontsize=12, pad=35)

            caminho_img_pizza = f'temp_grafico_pizza_{len(nomes_abas_usados)}.png'
            fig_pizza.savefig(caminho_img_pizza, dpi=120, bbox_inches='tight', pad_inches=0.2)
            plt.close(fig_pizza)
            imagens_temporarias.append(caminho_img_pizza)
            print(f"Gráfico de pizza salvo em: {caminho_img_pizza}")

            # =========================
            # 4. GRÁFICO DE BARRAS AUTOMÁTICO
            # =========================
            print("Criando gráfico de barras...")

            categorias_originais = [str(i) for i in dados_grafico.index]
            valores = dados_grafico['%'].values

            maior_legenda = max(len(c) for c in categorias_originais) if categorias_originais else 0
            quantidade_categorias = len(categorias_originais)

            usar_horizontal = (maior_legenda > 20) or (quantidade_categorias > 10)

            if usar_horizontal:
                print("Modo escolhido: gráfico de barras horizontal")

                fig_barra, ax_barra = plt.subplots(figsize=(16, max(8, quantidade_categorias * 0.6)))

                categorias = [textwrap.fill(c, width=35) for c in categorias_originais]
                y = np.arange(len(categorias))

                barras = ax_barra.barh(y, valores, height=0.55)

                ax_barra.set_title(titulo_formatado, fontsize=12, pad=20)
                ax_barra.set_xlabel("Porcentagem (%)")
                ax_barra.set_ylabel("Respostas")
                ax_barra.set_xlim(0, 100)

                ax_barra.set_yticks(y)
                ax_barra.set_yticklabels(categorias)
                ax_barra.invert_yaxis()

                for barra in barras:
                    largura = barra.get_width()
                    ax_barra.annotate(
                        f'{largura:.1f}%',
                        xy=(largura, barra.get_y() + barra.get_height() / 2),
                        xytext=(4, 0),
                        textcoords="offset points",
                        ha='left',
                        va='center',
                        fontsize=8
                    )

            else:
                print("Modo escolhido: gráfico de barras vertical")

                fig_barra, ax_barra = plt.subplots(figsize=(16, 8))

                categorias = [textwrap.fill(c, width=20) for c in categorias_originais]
                x = np.arange(len(categorias)) * 0.35
                barras = ax_barra.bar(x, valores, width=0.23)

                ax_barra.set_title(titulo_formatado, fontsize=12, pad=20)
                ax_barra.set_ylabel("Porcentagem (%)")
                ax_barra.set_xlabel("Respostas")
                ax_barra.set_ylim(0, 100)

                ax_barra.set_xticks(x)
                ax_barra.set_xticklabels(categorias, rotation=45, ha='right')

                for barra in barras:
                    altura = barra.get_height()
                    ax_barra.annotate(
                        f'{altura:.1f}%',
                        xy=(barra.get_x() + barra.get_width() / 2, altura),
                        xytext=(0, 3),
                        textcoords="offset points",
                        ha='center',
                        va='bottom',
                        fontsize=8
                    )

            fig_barra.tight_layout()

            caminho_img_barra = f'temp_grafico_barra_{len(nomes_abas_usados)}.png'
            fig_barra.savefig(caminho_img_barra, dpi=120, bbox_inches='tight', pad_inches=0.2)
            plt.close(fig_barra)
            imagens_temporarias.append(caminho_img_barra)
            print(f"Gráfico de barras salvo em: {caminho_img_barra}")

            # =========================
            # 5. INSERIR IMAGENS NA PLANILHA
            # =========================
            worksheet = writer.sheets[novo_nome_aba]

            img_pizza = Image(caminho_img_pizza)
            img_barra = Image(caminho_img_barra)

            # Coluna inicial para imagens
            coluna_imagem_idx = coluna_inicio_resumo + 4

            # Pizza
            celula_imagem_pizza = f"{get_column_letter(coluna_imagem_idx)}2"
            worksheet.add_image(img_pizza, celula_imagem_pizza)

            # Barras abaixo da pizza
            celula_imagem_barra = f"{get_column_letter(coluna_imagem_idx)}45"
            worksheet.add_image(img_barra, celula_imagem_barra)

            print(f"Imagens inseridas na aba '{novo_nome_aba}'.")

            # Ajustar largura das colunas do resumo
            worksheet.column_dimensions[get_column_letter(coluna_inicio_resumo + 1)].width = 45
            worksheet.column_dimensions[get_column_letter(coluna_inicio_resumo + 2)].width = 12
            worksheet.column_dimensions[get_column_letter(coluna_inicio_resumo + 3)].width = 10

    # 6. Limpeza final
    print("\nArquivos temporários gerados:")
    for img_path in imagens_temporarias:
        print(img_path, os.path.exists(img_path))

    for img_path in imagens_temporarias:
        if os.path.exists(img_path):
            os.remove(img_path)

    print(f"\nSucesso! O arquivo '{caminho_arquivo_saida}' foi gerado com todas as abas, tabelas e gráficos configurados.")

except Exception as e:
    print(f"Ocorreu um erro inesperado: {e}")